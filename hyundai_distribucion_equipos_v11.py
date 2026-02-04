from __future__ import annotations


# -*- coding: utf-8 -*-
"""
HYUNDAI - Distribución a Equipos (módulo independiente)
======================================================

Qué hace:
- Registrar distribución de diésel a equipos desde el HYUNDAI:
  (1) Con horómetro
  (2) Sin horómetro
- Listar registros (con filtros opcionales por fecha / equipo / responsable)
- Editar o eliminar cualquier registro por ID
- Exportar reporte mensual a Excel (solo hoja "Distribucion_Equipos") usando una plantilla .xlsx

Nuevo requisito implementado:
- Al registrar:
  1) Pide primero: "Litros despachados al equipo"
  2) Calcula automáticamente Contador Final = Contador Inicial + Litros
  3) Solo pide confirmar o cambiar el Contador Final antes de guardar

Dependencias:
    pip install pandas openpyxl

Archivos (por defecto):
- Base de datos SQLite: control_diesel.db
- Plantilla Excel: "Distribución de diésel a equipos desde el HYUNDAI_reporte_2025-09.xlsx"

Notas:
- Este script NO modifica recepción P2 ni despacho a HYUNDAI; solo trabaja la tabla
  "distribucion_hyundai_equipos".
- Incluye una migración ligera: agrega columnas faltantes si tu tabla es antigua.
"""
import os
import re
import sqlite3
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import openpyxl
from copy import copy


def parse_fecha_flexible(s: str) -> str:
    """
    Acepta:
      - '' -> hoy
      - 'DD' o 'D' -> usa año y mes actuales
      - 'YYYY-MM-DD' -> fecha completa
    Devuelve 'YYYY-MM-DD'. Lanza ValueError si es inválida.
    """
    s = (s or "").strip()
    hoy = datetime.now()
    if s == "":
        return hoy.strftime("%Y-%m-%d")
    if re.fullmatch(r"\d{1,2}", s):
        dia = int(s)
        y, m = hoy.year, hoy.month
        # último día del mes
        if m == 12:
            next_month = datetime(y + 1, 1, 1)
        else:
            next_month = datetime(y, m + 1, 1)
        last_day = (next_month - timedelta(days=1)).day
        if not (1 <= dia <= last_day):
            raise ValueError(f"Día inválido para el mes actual: {dia} (1..{last_day})")
        return f"{y:04d}-{m:02d}-{dia:02d}"
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", s):
        datetime.strptime(s, "%Y-%m-%d")
        return s
    raise ValueError("Formato inválido. Use YYYY-MM-DD o solo el día (DD).")

DB_PATH = "control_diesel.db"
TEMPLATE_XLSX = "Distribución de diésel a equipos desde el HYUNDAI_reporte_2025-09.xlsx"


# =========================
# Catálogos (ajusta a gusto)
# =========================
EQUIPOS = [
    "GENERADOR P1 CHICO", "GENERADOR P2", "GENERADOR P1", "BLOWER CISTERNA MOVIL", "CARGADOR SEM 636D", "CAMION AH7922",
    "CARGADOR SEM 639C", "LIMPIEZA P1", "LIMPIEZA P2", "PICK UP CE9798", "PALA DOOSAN", "RETRO JCB RE02", "TORRE DE LUZ RL4000", "CAMION AU8648",
    "CAMION KODIAK AB7045", "PICK UP 967767 AMADOR MANYOMA", "LIMPIEZA BOMBA DE TRASIEGO", "LIMPIEZA VAGON  DE CAMION EP7333",
    "CAMIONSITO DEL DIESEL AB7045", "DISTRIBUIDORA AH1509", "LIMPIEZA DE FILTROS CAMION AU8648", "TANQUE GRIS P1"
]

EQUIPOS_SIN_HOROMETRO = [
    "BLOWER CISTERNA MOVIL", "CAMION AH7922", "CARGADOR SEM 639C", "LIMPIEZA P1", "LIMPIEZA P2", "PICK UP CE9798",
    "TORRE DE LUZ RL4000", "CAMION AU8648", "CAMION KODIAK AB7045", "PICK UP 967767 AMADOR MANYOMA", "LIMPIEZA BOMBA DE TRASIEGO",
    "LIMPIEZA VAGON  DE CAMION EP7333", "CAMIONSITO DEL DIESEL AB7045", "DISTRIBUIDORA AH1509", "LIMPIEZA DE FILTROS CAMION AU8648", "TANQUE GRIS P1"
]

RESPONSABLES = ["Allan", "Alexander", "Manioma", "Jose", "Banega"]


# =========================
# Helpers de consola
# =========================
def pedir_float(msg: str, allow_empty: bool = False, default: Optional[float] = None) -> Optional[float]:
    while True:
        s = input(msg).strip()
        if s == "":
            if allow_empty:
                return default
            print("⚠️ No puede estar vacío.")
            continue
        try:
            return float(s)
        except ValueError:
            print("⚠️ Ingrese un número válido.")


def pedir_float_nav(msg: str, allow_empty: bool = False, default: float | None = None):
    while True:
        s = nav_input(msg)
        if s in (NAV_BACK, NAV_CANCEL):
            return s
        if s == "":
            if allow_empty:
                return default
            print("⚠️ No puede estar vacío.")
            continue
        try:
            return float(s)
        except ValueError:
            print("⚠️ Ingrese un número válido.")

def pedir_texto_nav(msg: str, allow_empty: bool = False, default: str | None = None):
    while True:
        s = nav_input(msg)
        if s in (NAV_BACK, NAV_CANCEL):
            return s
        if s == "":
            if allow_empty:
                return default
            print("⚠️ No puede estar vacío.")
            continue
        return s
def pedir_texto(msg: str, allow_empty: bool = False, default: Optional[str] = None) -> Optional[str]:
    s = input(msg).strip()
    if s == "":
        return default if allow_empty else None
    return s


# =========================
# Navegación (retroceso)
# =========================
NAV_BACK = "__BACK__"
NAV_CANCEL = "__CANCEL__"

def _nav_hint() -> str:
    return " (B=retroceder, Q=cancelar)"

def nav_input(prompt: str) -> str:
    s = input(prompt + _nav_hint() + ": ").strip()
    if s.upper() == "B":
        return NAV_BACK
    if s.upper() == "Q":
        return NAV_CANCEL
    return s
def interpretar_fecha(prompt: str = "📅 Fecha (YYYY-MM-DD) [Enter = hoy]: ") -> str:
    entrada = input(prompt).strip()
    hoy = datetime.now().strftime("%Y-%m-%d")
    if entrada == "":
        return hoy
    if entrada.isdigit() and 1 <= int(entrada) <= 31:
        return datetime.now().strftime(f"%Y-%m-{int(entrada):02d}")
    try:
        datetime.strptime(entrada, "%Y-%m-%d")
        return entrada
    except ValueError:
        print("⚠️ Fecha inválida. Usando hoy.")
        return hoy

def interpretar_hora(prompt: str = "🕒 Hora (HH:MM) [Enter = ahora]: ") -> str:
    entrada = input(prompt).strip()
    ahora = datetime.now().strftime("%H:%M")
    if entrada == "":
        return ahora
    try:
        datetime.strptime(entrada, "%H:%M")
        return entrada
    except ValueError:
        print("⚠️ Hora inválida. Usando ahora.")
        return ahora

def seleccionar_lista(nombre: str, opciones: List[str]) -> str:
    print(f"\nSeleccione {nombre}:")
    for i, op in enumerate(opciones, 1):
        print(f"  {i}. {op}")
    while True:
        s = input("Número: ").strip()
        if s.isdigit():
            idx = int(s)
            if 1 <= idx <= len(opciones):
                return opciones[idx - 1]
        print("⚠️ Selección inválida.")



def seleccionar_lista_nav(label: str, opciones: list[str]):
    """Selector con retroceso/cancelación. Devuelve opción elegida o NAV_BACK/NAV_CANCEL."""
    print(f"\nSeleccione {label}:")
    print("  B. ⬅️  Retroceder")
    print("  Q. ✖️  Cancelar")
    for i, op in enumerate(opciones, 1):
        print(f"  {i}. {op}")
    while True:
        s = input("Número (B retroceder / Q cancelar): ").strip()
        if s.upper() == "B":
            return NAV_BACK
        if s.upper() == "Q":
            return NAV_CANCEL
        if s.isdigit():
            idx = int(s)
            if 1 <= idx <= len(opciones):
                return opciones[idx - 1]
        print("⚠️ Selección inválida.")
def seleccionar_equipo(opciones: List[str]) -> str:
    """
    Selecciona un equipo de la lista, pero permite agregar uno manualmente.
    - Opción 0: escribir equipo manualmente
    Devuelve el nombre en MAYÚSCULA.
    """
    print("\nSeleccione equipo:")
    print("  0. ✍️  Escribir equipo manualmente")
    for i, op in enumerate(opciones, 1):
        print(f"  {i}. {op}")
    while True:
        s = input("Número (0 para manual): ").strip()
        if s == "0":
            manual = input("🛠️ Escribe el nombre del equipo: ").strip()
            if manual:
                return manual.upper()
            print("⚠️ El nombre no puede estar vacío.")
            continue
        if s.isdigit():
            idx = int(s)
            if 1 <= idx <= len(opciones):
                return str(opciones[idx - 1]).strip().upper()
        print("⚠️ Selección inválida.")



def seleccionar_equipo_nav(opciones: list[str]):
    """Como seleccionar_equipo, pero permite retroceder/cancelar."""
    print("\nSeleccione equipo:")
    print("  B. ⬅️  Retroceder")
    print("  Q. ✖️  Cancelar")
    print("  0. ✍️  Escribir equipo manualmente")
    for i, op in enumerate(opciones, 1):
        print(f"  {i}. {op}")
    while True:
        s = input("Número (0 manual / B retroceder / Q cancelar): ").strip()
        if s.upper() == "B":
            return NAV_BACK
        if s.upper() == "Q":
            return NAV_CANCEL
        if s == "0":
            manual = input("🛠️ Escribe el nombre del equipo: ").strip()
            if manual:
                return manual.upper()
            print("⚠️ El nombre no puede estar vacío.")
            continue
        if s.isdigit():
            idx = int(s)
            if 1 <= idx <= len(opciones):
                return str(opciones[idx - 1]).strip().upper()
        print("⚠️ Selección inválida.")
def litros_a_gal(litros: float) -> float:
    return round(litros * 0.264172, 2)
# =========================
# Secuencia del contador (derivada)
# =========================
def aplicar_secuencia_contador(df: pd.DataFrame) -> pd.DataFrame:
    """
    Construye la secuencia del contador cuando "contador_inicial" / "contador_final"
    vienen en blanco (NULL) en la base de datos.

    Reglas:
      - Ordena por fecha, hora, id (si existen)
      - Para cada fila:
          contador_inicial_calc = contador_inicial si existe; si no, toma el contador_final_calc anterior (o 0)
          contador_final_calc   = contador_final si existe; si no, contador_inicial_calc + litros_despachados
      - Crea columnas extra:
          Secuencia_Contador = "CI → CF"
          Delta_Litros = CF - CI
    """
    if df is None or df.empty:
        return df

    for c in ["contador_inicial", "contador_final", "litros_despachados"]:
        if c not in df.columns:
            df[c] = None

    sort_cols = [c for c in ["fecha", "hora", "id"] if c in df.columns]
    if sort_cols:
        df = df.sort_values(sort_cols).reset_index(drop=True)

    ci_calc = []
    cf_calc = []
    prev_final = None

    for _, r in df.iterrows():
        litros = r.get('litros_despachados')
        try:
            litros_val = float(litros)
        except Exception:
            litros_val = 0.0

        ci = r.get('contador_inicial')
        cf = r.get('contador_final')

        ci_valid = None
        cf_valid = None

        try:
            if ci is not None and str(ci) != 'nan':
                ci_valid = float(ci)
        except Exception:
            ci_valid = None

        try:
            if cf is not None and str(cf) != 'nan':
                cf_valid = float(cf)
        except Exception:
            cf_valid = None

        if ci_valid is None:
            ci_valid = 0.0 if prev_final is None else prev_final

        if cf_valid is None:
            cf_valid = round(ci_valid + litros_val, 2)

        prev_final = cf_valid

        ci_calc.append(round(ci_valid, 2))
        cf_calc.append(round(cf_valid, 2))

    df['contador_inicial_calc'] = ci_calc
    df['contador_final_calc'] = cf_calc

    # Preferir valores reales en BD si existen; si están vacíos, usar los calculados
    if 'contador_inicial' in df.columns:
        df['contador_inicial_show'] = df['contador_inicial']
    else:
        df['contador_inicial_show'] = None

    if 'contador_final' in df.columns:
        df['contador_final_show'] = df['contador_final']
    else:
        df['contador_final_show'] = None

    # Si vienen NaN/NULL, reemplazar por *_calc
    df['contador_inicial_show'] = df['contador_inicial_show'].where(~df['contador_inicial_show'].isna(), df['contador_inicial_calc'])
    df['contador_final_show'] = df['contador_final_show'].where(~df['contador_final_show'].isna(), df['contador_final_calc'])

    df['Delta_Litros'] = (df['contador_final_show'] - df['contador_inicial_show']).round(2)
    df['Secuencia_Contador'] = df['contador_inicial_show'].map(lambda x: f"{float(x):.2f}") + ' → ' + df['contador_final_show'].map(lambda x: f"{float(x):.2f}")
    return df



# =========================
# DB helpers
# =========================
def connect(db_path: str) -> sqlite3.Connection:
    return sqlite3.connect(db_path)

def table_exists(conn: sqlite3.Connection, table: str) -> bool:
    cur = conn.cursor()
    cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table,))
    return cur.fetchone() is not None

def get_columns(conn: sqlite3.Connection, table: str) -> List[str]:
    cur = conn.cursor()
    cur.execute(f"PRAGMA table_info({table})")
    return [r[1] for r in cur.fetchall()]

def column_map(conn: sqlite3.Connection, table: str) -> Dict[str, str]:
    """Mapa case-insensitive: lower_name -> actual_name en SQLite."""
    cols = get_columns(conn, table)
    return {c.lower(): c for c in cols}

def normalize_df_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Normaliza nombres de columnas comunes para evitar problemas de mayúsculas/minúsculas."""
    if df is None or df.empty:
        return df
    rename = {}
    for c in df.columns:
        cl = str(c).lower()
        # Contadores (en tu DB existen como Contador_inicial / Contador_final)
        if cl == "contador_inicial":
            rename[c] = "contador_inicial"
        elif cl == "contador_final":
            rename[c] = "contador_final"
    if rename:
        df = df.rename(columns=rename)
    return df


def ensure_schema(db_path: str) -> None:
    """
    Asegura que exista la tabla distribucion_hyundai_equipos con columnas necesarias.
    - Si la tabla ya existe, agrega columnas faltantes.
    - Si no existe, la crea con un esquema compatible.
    """
    with connect(db_path) as conn:
        cur = conn.cursor()

        if not table_exists(conn, "distribucion_hyundai_equipos"):
            cur.execute("""
                CREATE TABLE distribucion_hyundai_equipos (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    fecha TEXT,
                    hora TEXT,
                    equipo TEXT,
                    volumen_despachado REAL,        -- gal
                    responsable TEXT,

                    litros_despachados REAL,        -- litros
                    horometro_inicial REAL,
                    horometro_final REAL,
                    horas_trabajadas REAL,
                    consumo_por_gl_h REAL,          -- gal/h

                    precio_diesel REAL,             -- usd/gal
                    costo_diesel_usd REAL,          -- usd
                    contador_inicial REAL,          -- litros (contador/meter)
                    contador_final REAL,            -- litros
                    volumen_restante_hyundai REAL,  -- gal (si lo usas)
                    tipo_registro TEXT              -- HOROMETRO / SIN_HOROMETRO
                )
            """)
            conn.commit()
            return

        cols = set(get_columns(conn, "distribucion_hyundai_equipos"))
        desired = [
            ("hora", "TEXT"),
            ("litros_despachados", "REAL"),
            ("horometro_inicial", "REAL"),
            ("horometro_final", "REAL"),
            ("horas_trabajadas", "REAL"),
            ("consumo_por_gl_h", "REAL"),
            ("precio_diesel", "REAL"),
            ("costo_diesel_usd", "REAL"),
            ("contador_inicial", "REAL"),
            ("contador_final", "REAL"),
            ("volumen_restante_hyundai", "REAL"),
            ("tipo_registro", "TEXT"),
        ]

        for name, typ in desired:
            if name not in cols:
                try:
                    cur.execute(f"ALTER TABLE distribucion_hyundai_equipos ADD COLUMN {name} {typ}")
                except Exception:
                    pass
        conn.commit()

def fetch_last_contador_final(conn: sqlite3.Connection) -> float:
    """Último contador_final GLOBAL registrado."""
    cols = set(get_columns(conn, "distribucion_hyundai_equipos"))
    cur = conn.cursor()
    if "fecha" in cols and "hora" in cols:
        cur.execute("""
            SELECT contador_final
            FROM distribucion_hyundai_equipos
            WHERE contador_final IS NOT NULL
            ORDER BY fecha DESC, hora DESC, id DESC
            LIMIT 1
        """)
    else:
        cur.execute("""
            SELECT contador_final
            FROM distribucion_hyundai_equipos
            WHERE contador_final IS NOT NULL
            ORDER BY id DESC
            LIMIT 1
        """)
    row = cur.fetchone()
    return round(float(row[0]), 2) if row and row[0] is not None else 0.0

def fetch_last_horometro_final(conn: sqlite3.Connection, equipo: str) -> float:
    """Último horometro_final del equipo (para horómetro inicial por defecto)."""
    cols = set(get_columns(conn, "distribucion_hyundai_equipos"))
    cur = conn.cursor()
    if "fecha" in cols and "hora" in cols:
        cur.execute("""
            SELECT horometro_final
            FROM distribucion_hyundai_equipos
            WHERE equipo = ? AND horometro_final IS NOT NULL
            ORDER BY fecha DESC, hora DESC, id DESC
            LIMIT 1
        """, (equipo,))
    else:
        cur.execute("""
            SELECT horometro_final
            FROM distribucion_hyundai_equipos
            WHERE equipo = ? AND horometro_final IS NOT NULL
            ORDER BY id DESC
            LIMIT 1
        """, (equipo,))
    row = cur.fetchone()
    return round(float(row[0]), 2) if row and row[0] is not None else 0.0

def insert_distribucion(conn: sqlite3.Connection, data: Dict[str, Any]) -> int:
    cmap = column_map(conn, "distribucion_hyundai_equipos")
    # mapear keys a nombre real de la BD (case-insensitive)
    data2 = {}
    for k, v in data.items():
        kk = str(k).lower()
        if kk in cmap:
            data2[cmap[kk]] = v
    keys = list(data2.keys())
    placeholders = ", ".join(["?"] * len(keys))
    sql = f"INSERT INTO distribucion_hyundai_equipos ({', '.join(keys)}) VALUES ({placeholders})"
    cur = conn.cursor()
    cur.execute(sql, tuple(data2[k] for k in keys))
    conn.commit()
    return int(cur.lastrowid)

def update_distribucion(conn: sqlite3.Connection, record_id: int, data: Dict[str, Any]) -> None:
    cmap = column_map(conn, "distribucion_hyundai_equipos")
    data2 = {}
    for k, v in data.items():
        kk = str(k).lower()
        if kk in cmap:
            data2[cmap[kk]] = v
    if not data2:
        return
    sets = ", ".join([f"{k}=?" for k in data2.keys()])
    sql = f"UPDATE distribucion_hyundai_equipos SET {sets} WHERE id=?"
    cur = conn.cursor()
    cur.execute(sql, tuple(data2.values()) + (record_id,))
    conn.commit()

def delete_distribucion(conn: sqlite3.Connection, record_id: int) -> bool:
    cur = conn.cursor()
    cur.execute("DELETE FROM distribucion_hyundai_equipos WHERE id=?", (record_id,))
    conn.commit()
    return cur.rowcount > 0

def fetch_by_id(conn: sqlite3.Connection, record_id: int) -> Optional[Dict[str, Any]]:
    cur = conn.cursor()
    cols = get_columns(conn, "distribucion_hyundai_equipos")
    cur.execute(f"SELECT {', '.join(cols)} FROM distribucion_hyundai_equipos WHERE id=?", (record_id,))
    row = cur.fetchone()
    if not row:
        return None
    rec = dict(zip(cols, row))
    # alias case-insensitive contadores
    for k in list(rec.keys()):
        kl = str(k).lower()
        if kl == 'contador_inicial':
            rec['contador_inicial'] = rec[k]
        if kl == 'contador_final':
            rec['contador_final'] = rec[k]
    return rec


# =========================
# Precio diésel (compat)
# =========================
def obtener_precio_diesel_actual() -> float:
    """
    Mantiene compatibilidad: intenta leer de reportes_DMI.db o reportes.db.
    Si no existe o falla, devuelve 0.0.
    """
    consulta = "SELECT precio_diesel FROM {} WHERE precio_diesel IS NOT NULL ORDER BY fecha_produccion DESC LIMIT 1"
    for db_name, table in [("reportes_DMI.db", "reportes_DMI"), ("reportes.db", "reportes")]:
        if not os.path.exists(db_name):
            continue
        try:
            conn = sqlite3.connect(db_name)
            cur = conn.cursor()
            cur.execute(consulta.format(table))
            row = cur.fetchone()
            conn.close()
            if row and row[0] is not None:
                return float(row[0])
        except Exception:
            pass
    return 2.7955


# =========================
# Flujo nuevo contador auto
# =========================
def pedir_contadores_auto(conn: sqlite3.Connection, litros_despachados: float) -> Tuple[float, float]:
    contador_inicial_default = fetch_last_contador_final(conn)

    entrada_ci = pedir_texto(f"📟 Contador inicial (litros) [{contador_inicial_default}]: ",
                            allow_empty=True, default=str(contador_inicial_default))
    try:
        contador_inicial = float(entrada_ci) if entrada_ci is not None else contador_inicial_default
    except ValueError:
        print("⚠️ Contador inicial inválido. Usando el valor por defecto.")
        contador_inicial = contador_inicial_default

    contador_final_calc = round(contador_inicial + litros_despachados, 2)

    entrada_cf = pedir_texto(
        f"📟 Contador final calculado = {contador_final_calc} (Enter para aceptar / escribe otro): ",
        allow_empty=True,
        default=str(contador_final_calc),
    )
    try:
        contador_final = float(entrada_cf) if entrada_cf is not None else contador_final_calc
    except ValueError:
        print("⚠️ Contador final inválido. Se usará el calculado.")
        contador_final = contador_final_calc

    return round(contador_inicial, 2), round(contador_final, 2)


# =========================
# Registrar
# =========================

# =========================
# Backfill: guardar secuencia de contador en BD
# =========================
def backfill_contadores(db_path: str) -> None:
    """
    Llena (solo donde están NULL) los campos contador_inicial y contador_final en la BD,
    siguiendo la secuencia en el ORDEN REAL de la base:
      ORDER BY fecha ASC, hora ASC, id ASC

    Importante:
    - Si un registro YA tiene contador_inicial/contador_final, se respetan.
    - Si faltan, se calculan usando:
        contador_inicial = contador_final del registro anterior
        contador_final   = contador_inicial + litros_despachados
    """
    ensure_schema(db_path)
    with connect(db_path) as conn:
        cols = get_columns(conn, "distribucion_hyundai_equipos")
        if "litros_despachados" not in cols:
            print("⚠️ Tu tabla no tiene 'litros_despachados'. No puedo reconstruir contadores.")
            return
        cmap = column_map(conn, 'distribucion_hyundai_equipos')
        if 'contador_inicial' not in cmap or 'contador_final' not in cmap:
            print("⚠️ Tu tabla no tiene columnas de contador.")
            return

        df = pd.read_sql_query(
            f"SELECT {', '.join(cols)} FROM distribucion_hyundai_equipos ORDER BY fecha ASC, hora ASC, id ASC",
            conn
        )
        df = normalize_df_columns(df)
        if df.empty:
            print("📭 No hay registros.")
            return

        df = aplicar_secuencia_contador(df)

        # Actualizar solo filas donde contador_inicial o contador_final están NULL
        cur = conn.cursor()
        updated_rows = 0
        for _, r in df.iterrows():
            rid = int(r["id"])
            ci_db = r.get("contador_inicial")
            cf_db = r.get("contador_final")

            needs = (ci_db is None) or (cf_db is None) or (str(ci_db) == "nan") or (str(cf_db) == "nan")
            if not needs:
                continue

            ci = float(r["contador_inicial_calc"])
            cf = float(r["contador_final_calc"])
            cur.execute(
                f"UPDATE distribucion_hyundai_equipos SET {cmap['contador_inicial']}=?, {cmap['contador_final']}=? WHERE id=?",
                (ci, cf, rid)
            )
            updated_rows += 1

        conn.commit()
        print(f"✅ Backfill completo. Registros actualizados: {updated_rows}")

def registrar_con_horometro(db_path: str) -> None:
    ensure_schema(db_path)
    with connect(db_path) as conn:
        print("\n🛠️ REGISTRAR DISTRIBUCIÓN (CON HORÓMETRO)")
        state: dict = {}

        steps = ["fecha","hora","equipo","responsable","litros","contador","horometro","precio"]
        i = 0

        while i < len(steps):
            step = steps[i]

            if step == "fecha":
                val = pedir_texto_nav("📅 Fecha (YYYY-MM-DD) o solo día (DD) [Enter = hoy]", allow_empty=True, default="")
                if val == NAV_CANCEL:
                    print("Cancelado."); return
                if val == NAV_BACK:
                    i = max(i-1, 0); continue
                try:
                    state["fecha"] = parse_fecha_flexible(val)
                except Exception as e:
                    print(f"⚠️ {e}")
                    continue
                i += 1; continue

            if step == "hora":
                val = pedir_texto_nav("🕒 Hora (HH:MM) [Enter = ahora]", allow_empty=True, default="")
                if val == NAV_CANCEL:
                    print("Cancelado."); return
                if val == NAV_BACK:
                    i = max(i-1, 0); continue
                state["hora"] = val if val else datetime.now().strftime("%H:%M")
                i += 1; continue

            if step == "equipo":
                val = seleccionar_equipo_nav(EQUIPOS)
                if val == NAV_CANCEL:
                    print("Cancelado."); return
                if val == NAV_BACK:
                    i = max(i-1, 0); continue
                state["equipo"] = val
                i += 1; continue

            if step == "responsable":
                state["responsable"] = seleccionar_lista("responsable", RESPONSABLES)
                i += 1; continue

            if step == "litros":
                val = pedir_float_nav("⛽ Litros despachados al equipo")
                if val == NAV_CANCEL:
                    print("Cancelado."); return
                if val == NAV_BACK:
                    i = max(i-1, 0); continue
                litros = float(val)
                state["litros_despachados"] = litros
                state["volumen_despachado_gal"] = litros_a_gal(litros)
                print(f"🔁 Conversión: {litros} litros = {state['volumen_despachado_gal']} galones")
                i += 1; continue

            if step == "contador":
                contador_inicial_default = fetch_last_contador_final(conn)
                ci = pedir_texto_nav(f"📟 Contador inicial (litros) [{contador_inicial_default}]", allow_empty=True, default=str(contador_inicial_default))
                if ci == NAV_CANCEL:
                    print("Cancelado."); return
                if ci == NAV_BACK:
                    i = max(i-1, 0); continue
                try:
                    contador_inicial = float(ci)
                except Exception:
                    contador_inicial = float(contador_inicial_default)

                contador_final_calc = round(contador_inicial + float(state.get("litros_despachados", 0.0)), 2)
                cf = pedir_texto_nav(f"📟 Contador final calculado = {contador_final_calc} (Enter para aceptar / escribe otro)", allow_empty=True, default=str(contador_final_calc))
                if cf == NAV_CANCEL:
                    print("Cancelado."); return
                if cf == NAV_BACK:
                    i = max(i-1, 0); continue
                try:
                    contador_final = float(cf)
                except Exception:
                    contador_final = float(contador_final_calc)

                state["contador_inicial"] = round(contador_inicial, 2)
                state["contador_final"] = round(contador_final, 2)
                i += 1; continue

            if step == "horometro":
                hi_default = fetch_last_horometro_final(conn, state["equipo"])
                hi = pedir_texto_nav(f"⏱️ Horómetro inicial [{hi_default}]", allow_empty=True, default=str(hi_default))
                if hi == NAV_CANCEL:
                    print("Cancelado."); return
                if hi == NAV_BACK:
                    i = max(i-1, 0); continue
                try:
                    horometro_inicial = float(hi)
                except Exception:
                    horometro_inicial = float(hi_default)

                hf = pedir_float_nav("⏱️ Horómetro final (valor actual del equipo)")
                if hf == NAV_CANCEL:
                    print("Cancelado."); return
                if hf == NAV_BACK:
                    i = max(i-1, 0); continue
                horometro_final = float(hf)

                horas = round(horometro_final - horometro_inicial, 2)
                if horas <= 0:
                    print("⚠️ Horas trabajadas no pueden ser <= 0. Usa B para retroceder.")
                    continue

                consumo = round(float(state["volumen_despachado_gal"]) / horas, 2)

                state["horometro_inicial"] = horometro_inicial
                state["horometro_final"] = horometro_final
                state["horas_trabajadas"] = horas
                state["consumo_por_gl_h"] = consumo
                i += 1; continue

            if step == "precio":
                precio = obtener_precio_diesel_actual()
                p = pedir_texto_nav(f"💲 Precio diésel actual = {precio} USD/gal. ¿Deseas cambiarlo? [Enter = No]", allow_empty=True, default="")
                if p == NAV_CANCEL:
                    print("Cancelado."); return
                if p == NAV_BACK:
                    i = max(i-1, 0); continue
                if p:
                    try:
                        precio = float(p)
                    except ValueError:
                        print("⚠️ Precio inválido. Se usará el valor por defecto.")
                state["precio_diesel"] = precio
                i += 1; continue

        costo = None
        if state.get("precio_diesel") is not None:
            costo = round(float(state["volumen_despachado_gal"]) * float(state["precio_diesel"]), 2)

        data = dict(
            fecha=state["fecha"],
            hora=state["hora"],
            equipo=state["equipo"],
            responsable=state["responsable"],
            litros_despachados=float(state["litros_despachados"]),
            volumen_despachado=float(state["volumen_despachado_gal"]),
            contador_inicial=float(state["contador_inicial"]),
            contador_final=float(state["contador_final"]),
            horometro_inicial=float(state["horometro_inicial"]),
            horometro_final=float(state["horometro_final"]),
            horas_trabajadas=float(state["horas_trabajadas"]),
            consumo_por_gl_h=float(state["consumo_por_gl_h"]),
            precio_diesel=float(state["precio_diesel"]) if state.get("precio_diesel") is not None else None,
            costo_diesel_usd=costo,
            volumen_restante_hyundai=None,
            tipo_registro="HOROMETRO",
        )
        new_id = insert_distribucion(conn, data)
        print(f"✅ Registro guardado. ID = {new_id}")
        print(f"📟 Contador inicial: {state['contador_inicial']} | Contador final: {state['contador_final']}")
        print(f"🕒 Horas trabajadas: {state['horas_trabajadas']} | Consumo: {state['consumo_por_gl_h']} gal/h")
        print(f"💰 Costo estimado: ${costo:.2f}" if costo is not None else "💰 Costo estimado: N/A")


def registrar_sin_horometro(db_path: str) -> None:
    ensure_schema(db_path)
    with connect(db_path) as conn:
        print("\n🛠️ REGISTRAR DISTRIBUCIÓN (SIN HORÓMETRO)")
        state: dict = {}

        steps = ["fecha","hora","equipo","responsable","litros","contador","precio"]
        i = 0

        while i < len(steps):
            step = steps[i]

            if step == "fecha":
                val = pedir_texto_nav("📅 Fecha (YYYY-MM-DD) o solo día (DD) [Enter = hoy]", allow_empty=True, default="")
                if val == NAV_CANCEL:
                    print("Cancelado."); return
                if val == NAV_BACK:
                    i = max(i-1, 0); continue
                try:
                    state["fecha"] = parse_fecha_flexible(val)
                except Exception as e:
                    print(f"⚠️ {e}")
                    continue
                i += 1; continue

            if step == "hora":
                val = pedir_texto_nav("🕒 Hora (HH:MM) [Enter = ahora]", allow_empty=True, default="")
                if val == NAV_CANCEL:
                    print("Cancelado."); return
                if val == NAV_BACK:
                    i = max(i-1, 0); continue
                state["hora"] = val if val else datetime.now().strftime("%H:%M")
                i += 1; continue

            if step == "equipo":
                val = seleccionar_equipo_nav(EQUIPOS)
                if val == NAV_CANCEL:
                    print("Cancelado."); return
                if val == NAV_BACK:
                    i = max(i-1, 0); continue
                state["equipo"] = val
                i += 1; continue

            if step == "responsable":
                val = seleccionar_lista_nav("responsable", RESPONSABLES)
                if val == NAV_CANCEL:
                    print("Cancelado."); return
                if val == NAV_BACK:
                    i = max(i-1, 0); continue
                state["responsable"] = val
                i += 1; continue

            if step == "litros":
                val = pedir_float_nav("⛽ Litros despachados al equipo")
                if val == NAV_CANCEL:
                    print("Cancelado."); return
                if val == NAV_BACK:
                    i = max(i-1, 0); continue
                litros = float(val)
                state["litros_despachados"] = litros
                state["volumen_despachado_gal"] = litros_a_gal(litros)
                print(f"🔁 Conversión: {litros} litros = {state['volumen_despachado_gal']} galones")
                i += 1; continue

            if step == "contador":
                # GLOBAL HYUNDAI: contador inicial por defecto = último contador_final en BD
                contador_inicial_default = fetch_last_contador_final(conn)
                ci = pedir_texto_nav(f"📟 Contador inicial (litros) [{contador_inicial_default}]", allow_empty=True, default=str(contador_inicial_default))
                if ci == NAV_CANCEL:
                    print("Cancelado."); return
                if ci == NAV_BACK:
                    i = max(i-1, 0); continue
                try:
                    contador_inicial = float(ci)
                except Exception:
                    contador_inicial = float(contador_inicial_default)

                contador_final_calc = round(contador_inicial + float(state.get("litros_despachados", 0.0)), 2)
                cf = pedir_texto_nav(f"📟 Contador final calculado = {contador_final_calc} (Enter para aceptar / escribe otro)", allow_empty=True, default=str(contador_final_calc))
                if cf == NAV_CANCEL:
                    print("Cancelado."); return
                if cf == NAV_BACK:
                    i = max(i-1, 0); continue
                try:
                    contador_final = float(cf)
                except Exception:
                    contador_final = float(contador_final_calc)

                state["contador_inicial"] = round(contador_inicial, 2)
                state["contador_final"] = round(contador_final, 2)
                i += 1; continue

            if step == "precio":
                precio = obtener_precio_diesel_actual()
                p = pedir_texto_nav(f"💲 Precio diésel actual = {precio} USD/gal. ¿Deseas cambiarlo? [Enter = No]", allow_empty=True, default="")
                if p == NAV_CANCEL:
                    print("Cancelado."); return
                if p == NAV_BACK:
                    i = max(i-1, 0); continue
                if p:
                    try:
                        precio = float(p)
                    except ValueError:
                        print("⚠️ Precio inválido. Se usará el valor por defecto.")
                state["precio_diesel"] = precio
                i += 1; continue

        costo = None
        if state.get("precio_diesel") is not None:
            costo = round(float(state["volumen_despachado_gal"]) * float(state["precio_diesel"]), 2)

        data = dict(
            fecha=state["fecha"],
            hora=state["hora"],
            equipo=state["equipo"],
            responsable=state["responsable"],
            litros_despachados=float(state["litros_despachados"]),
            volumen_despachado=float(state["volumen_despachado_gal"]),
            contador_inicial=float(state["contador_inicial"]),
            contador_final=float(state["contador_final"]),
            horometro_inicial=None,
            horometro_final=None,
            horas_trabajadas=None,
            consumo_por_gl_h=None,
            precio_diesel=float(state["precio_diesel"]) if state.get("precio_diesel") is not None else None,
            costo_diesel_usd=costo,
            volumen_restante_hyundai=None,
            tipo_registro="SIN_HOROMETRO",
        )
        new_id = insert_distribucion(conn, data)
        print(f"✅ Registro guardado. ID = {new_id}")
        print(f"📟 Contador inicial: {state['contador_inicial']} | Contador final: {state['contador_final']}")
        print(f"💰 Costo estimado: ${costo:.2f}" if costo is not None else "💰 Costo estimado: N/A")


def listar(db_path: str) -> None:
    ensure_schema(db_path)
    with connect(db_path) as conn:
        cols = get_columns(conn, "distribucion_hyundai_equipos")

        print("\n🔎 LISTAR REGISTROS (filtros opcionales)")
        desde = pedir_texto("Desde (YYYY-MM-DD) [Enter = sin filtro]: ", allow_empty=True, default="")
        hasta = pedir_texto("Hasta (YYYY-MM-DD) [Enter = sin filtro]: ", allow_empty=True, default="")
        equipo = pedir_texto("Equipo contiene... [Enter = todos]: ", allow_empty=True, default="")
        responsable = pedir_texto("Responsable contiene... [Enter = todos]: ", allow_empty=True, default="")

        where = []
        params: List[Any] = []

        if desde:
            where.append("fecha >= ?")
            params.append(desde)
        if hasta:
            where.append("fecha <= ?")
            params.append(hasta)
        if equipo:
            where.append("equipo LIKE ?")
            params.append(f"%{equipo.strip().upper()}%")
        if responsable:
            where.append("responsable LIKE ?")
            params.append(f"%{responsable.strip()}%")

        sql = f"SELECT {', '.join(cols)} FROM distribucion_hyundai_equipos"
        if where:
            sql += " WHERE " + " AND ".join(where)
        sql += " ORDER BY fecha DESC, hora DESC, id DESC"

        df = pd.read_sql_query(sql, conn, params=params)
        df = normalize_df_columns(df)

        # aplicar secuencia del contador (si viene en blanco en BD)
        df = aplicar_secuencia_contador(df)

        if df.empty:
            print("📭 No hay registros para esos filtros.")
            return

        # Mostrar columnas clave
        show_cols = [c for c in ["id","fecha","hora","equipo","litros_despachados","volumen_despachado","responsable","contador_inicial","contador_final","Secuencia_Contador","Delta_Litros","tipo_registro"] if c in df.columns]
        print(df[show_cols].to_string(index=False))


# =========================
# Editar / Eliminar
# =========================
def editar(db_path: str) -> None:
    ensure_schema(db_path)
    with connect(db_path) as conn:
        rid = pedir_texto("✏️ ID a editar: ")
        if not rid or not rid.isdigit():
            print("⚠️ ID inválido.")
            return
        rid_i = int(rid)
        rec = fetch_by_id(conn, rid_i)
        if not rec:
            print("⚠️ No existe un registro con ese ID.")
            return

        print("\nRegistro actual:")
        for k in ["id","fecha","hora","equipo","responsable","litros_despachados","volumen_despachado","horometro_inicial","horometro_final","horas_trabajadas","consumo_por_gl_h","contador_inicial","contador_final","precio_diesel","costo_diesel_usd","tipo_registro"]:
            if k in rec:
                print(f"  {k}: {rec.get(k)}")

        # Campos editables principales
        fecha = pedir_texto(f"Fecha [{rec.get('fecha','')}]: ", allow_empty=True, default=rec.get("fecha"))
        hora = pedir_texto(f"Hora [{rec.get('hora','')}]: ", allow_empty=True, default=rec.get("hora"))
        equipo = pedir_texto(f"Equipo [{rec.get('equipo','')}]: ", allow_empty=True, default=rec.get("equipo"))
        responsable = pedir_texto(f"Responsable [{rec.get('responsable','')}]: ", allow_empty=True, default=rec.get("responsable"))

        # Si cambia litros, recalcular gal y sugerir contador_final
        litros_old = rec.get("litros_despachados")
        litros_new = pedir_texto(f"Litros despachados [{litros_old}]: ", allow_empty=True, default=str(litros_old) if litros_old is not None else "")
        try:
            litros_val = float(litros_new) if litros_new not in (None, "") else float(litros_old) if litros_old is not None else None
        except ValueError:
            print("⚠️ Litros inválidos. Se mantiene el valor anterior.")
            litros_val = float(litros_old) if litros_old is not None else None

        volumen_gal = rec.get("volumen_despachado")
        if litros_val is not None:
            volumen_gal = litros_a_gal(litros_val)

        # contador
        ci_old = rec.get("contador_inicial") if rec.get("contador_inicial") is not None else 0.0
        ci_in = pedir_texto(f"Contador inicial [{ci_old}]: ", allow_empty=True, default=str(ci_old))
        try:
            ci_val = float(ci_in) if ci_in is not None else float(ci_old)
        except ValueError:
            ci_val = float(ci_old)

        cf_sugerido = round(ci_val + (litros_val or 0.0), 2)
        cf_old = rec.get("contador_final")
        cf_in = pedir_texto(f"Contador final sugerido={cf_sugerido} (Enter para aceptar / escribe otro) [actual={cf_old}]: ",
                            allow_empty=True, default=str(cf_sugerido))
        try:
            cf_val = float(cf_in) if cf_in is not None else cf_sugerido
        except ValueError:
            cf_val = cf_sugerido

        # horómetro (si existe)
        hi_old = rec.get("horometro_inicial")
        hf_old = rec.get("horometro_final")
        hi_in = pedir_texto(f"Horómetro inicial [{hi_old}]: ", allow_empty=True, default=str(hi_old) if hi_old is not None else "")
        hf_in = pedir_texto(f"Horómetro final [{hf_old}]: ", allow_empty=True, default=str(hf_old) if hf_old is not None else "")

        def parse_opt_float(s: Optional[str]) -> Optional[float]:
            if s is None or s == "":
                return None
            try:
                return float(s)
            except ValueError:
                return None

        hi_val = parse_opt_float(hi_in)
        hf_val = parse_opt_float(hf_in)

        horas = rec.get("horas_trabajadas")
        consumo = rec.get("consumo_por_gl_h")
        if hi_val is not None and hf_val is not None:
            horas = round(hf_val - hi_val, 2)
            if horas <= 0:
                print("⚠️ Horas trabajadas <= 0; se dejará como estaba.")
                horas = rec.get("horas_trabajadas")
            else:
                consumo = round((volumen_gal or 0.0) / horas, 2) if volumen_gal is not None else None

        # precio / costo
        precio_old = rec.get("precio_diesel")
        precio_in = pedir_texto(f"Precio diesel (USD/gal) [{precio_old}]: ", allow_empty=True, default=str(precio_old) if precio_old is not None else "")
        precio_val = parse_opt_float(precio_in)
        if precio_val is None and precio_old is not None:
            precio_val = float(precio_old)

        costo = rec.get("costo_diesel_usd")
        if precio_val is not None and volumen_gal is not None:
            costo = round(volumen_gal * precio_val, 2)

        tipo = pedir_texto(f"Tipo registro [{rec.get('tipo_registro')}]: ", allow_empty=True, default=rec.get("tipo_registro"))

        update_distribucion(conn, rid_i, dict(
            fecha=fecha,
            hora=hora,
            equipo=(equipo or "").strip().upper() if equipo else None,
            responsable=responsable,
            litros_despachados=litros_val,
            volumen_despachado=volumen_gal,
            contador_inicial=ci_val,
            contador_final=cf_val,
            horometro_inicial=hi_val,
            horometro_final=hf_val,
            horas_trabajadas=horas,
            consumo_por_gl_h=consumo,
            precio_diesel=precio_val,
            costo_diesel_usd=costo,
            tipo_registro=tipo,
        ))
        print("✅ Registro actualizado.")

def eliminar(db_path: str) -> None:
    ensure_schema(db_path)
    with connect(db_path) as conn:
        rid = pedir_texto("🗑️ ID a eliminar: ")
        if not rid or not rid.isdigit():
            print("⚠️ ID inválido.")
            return
        rid_i = int(rid)
        rec = fetch_by_id(conn, rid_i)
        if not rec:
            print("⚠️ No existe un registro con ese ID.")
            return
        print(f"Vas a eliminar ID={rid_i} | {rec.get('fecha')} | {rec.get('equipo')} | {rec.get('litros_despachados')} L | {rec.get('responsable')}")
        conf = input("Escribe SI para confirmar: ").strip().upper()
        if conf == "SI":
            ok = delete_distribucion(conn, rid_i)
            print("✅ Eliminado." if ok else "⚠️ No se pudo eliminar.")
        else:
            print("Cancelado.")


# =========================
# Exportar a Excel (plantilla)
# =========================
EXPORT_COLS = [
    ("fecha", "📅 Fecha"),
    ("equipo", "🛠️ Equipo"),
    ("volumen_despachado", "⛽ Volumen Despachado (gal)"),
    ("responsable", "👤 Responsable"),
    ("litros_despachados", "🧪 Litros Despachados"),
    ("horas_trabajadas", "🕒 Horas Trabajadas"),
    ("horometro_inicial", "⏱️ Horómetro Inicial"),
    ("horometro_final", "⏱️ Horómetro Final"),
    ("contador_inicial", "📟 Contador Inicial"),
    ("contador_final", "📟 Contador Final"),
    ("consumo_por_gl_h", "🔁 Consumo (gal/h)"),
    ("precio_diesel", "💲 Precio Diesel"),
    ("costo_diesel_usd", "💰 Costo Diesel (USD)"),
]

def copy_row_style(src_ws, src_row: int, dst_ws, dst_row: int, max_col: int) -> None:
    for c in range(1, max_col + 1):
        sc = src_ws.cell(row=src_row, column=c)
        dc = dst_ws.cell(row=dst_row, column=c)
        if sc.has_style:
            dc._style = copy(sc._style)
        dc.number_format = sc.number_format
        dc.font = copy(sc.font)
        dc.border = copy(sc.border)
        dc.fill = copy(sc.fill)
        dc.alignment = copy(sc.alignment)
        dc.protection = copy(sc.protection)
        dc.comment = sc.comment

def copy_column_widths(src_ws, dst_ws, max_col: int) -> None:
    for c in range(1, max_col + 1):
        letter = openpyxl.utils.get_column_letter(c)
        if letter in src_ws.column_dimensions:
            dst_ws.column_dimensions[letter].width = src_ws.column_dimensions[letter].width

def exportar_mes(db_path: str, plantilla_xlsx: str) -> None:
    ensure_schema(db_path)
    mes = pedir_texto("📦 Mes a exportar (YYYY-MM) ej: 2025-09: ")
    if not mes or len(mes) != 7 or mes[4] != "-":
        print("⚠️ Formato de mes inválido.")
        return

    with connect(db_path) as conn:
        cols = get_columns(conn, "distribucion_hyundai_equipos")
        # filtro por mes usando LIKE en fecha (YYYY-MM-%)
        sql = f"SELECT {', '.join(cols)} FROM distribucion_hyundai_equipos WHERE fecha LIKE ? ORDER BY fecha ASC, hora ASC, id ASC"
        df = pd.read_sql_query(sql, conn, params=[f"{mes}-%"])
    df = normalize_df_columns(df)

    # aplicar secuencia del contador (si viene en blanco en BD)
    df = aplicar_secuencia_contador(df)

    if df.empty:
        print("📭 No hay registros para ese mes.")
        return

    # Preparar dataframe con columnas de exportación
    out = pd.DataFrame()
    for col_db, col_xl in EXPORT_COLS:
        if col_db in df.columns:
            out[col_xl] = df[col_db]
        else:
            out[col_xl] = None

    # Si BD no trae contador_inicial/contador_final (o vienen totalmente vacíos), usamos los *_calc ya calculados
    if '📟 Contador Inicial' in out.columns:
        if ('contador_inicial' not in df.columns) or (df['contador_inicial'].isna().all()):
            out['📟 Contador Inicial'] = df.get('contador_inicial_calc')
        else:
            out['📟 Contador Inicial'] = df.get('contador_inicial')

    if '📟 Contador Final' in out.columns:
        if ('contador_final' not in df.columns) or (df['contador_final'].isna().all()):
            out['📟 Contador Final'] = df.get('contador_final_calc')
        else:
            out['📟 Contador Final'] = df.get('contador_final')



    # Cargar plantilla
    if not os.path.exists(plantilla_xlsx):
        print(f"⚠️ No se encontró la plantilla: {plantilla_xlsx}")
        return

    wb = openpyxl.load_workbook(plantilla_xlsx)
    if "Distribucion_Equipos" not in wb.sheetnames:
        print("⚠️ La plantilla no tiene la hoja 'Distribucion_Equipos'.")
        return

    ws = wb["Distribucion_Equipos"]

    # Detectar # columnas por headers en fila 1
    max_col = ws.max_column

    # Guardar estilo de una fila de ejemplo (fila 2) si existe
    sample_row = 2 if ws.max_row >= 2 else 1

    # Limpiar datos existentes (de fila 2 en adelante)
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    # Escribir data desde fila 2
    start_row = 2
    for i, row in out.iterrows():
        r = start_row + i
        # Copiar estilos desde fila sample_row
        if sample_row >= 2:
            copy_row_style(ws, sample_row, ws, r, max_col)

        for j, (col_db, col_xl) in enumerate(EXPORT_COLS, start=1):
            ws.cell(row=r, column=j, value=row.get(col_xl))

    # Ajustar widths a lo mismo (ya está en plantilla, pero por si acaso)
    copy_column_widths(ws, ws, max_col)

    # Guardar
    out_name = f"Distribucion_Equipos_{mes}.xlsx"
    wb.save(out_name)
    print(f"✅ Exportado: {out_name}")


# =========================
# Menú
# =========================
def menu() -> None:
    while True:
        print("\n" + "="*60)
        print("HYUNDAI | Distribución a Equipos")
        print("="*60)
        print("1) Registrar distribución (CON horómetro)")
        print("2) Registrar distribución (SIN horómetro)")
        print("3) Listar registros")
        print("4) Editar registro por ID")
        print("5) Eliminar registro por ID")
        print("6) Exportar mes a Excel (solo hoja Distribucion_Equipos)")
        print("7) Reconstruir/llenar Contadores en BD (backfill)")
        print("0) Salir")
        op = input("Opción: ").strip()

        if op == "1":
            registrar_con_horometro(DB_PATH)
        elif op == "2":
            registrar_sin_horometro(DB_PATH)
        elif op == "3":
            listar(DB_PATH)
        elif op == "4":
            editar(DB_PATH)
        elif op == "5":
            eliminar(DB_PATH)
        elif op == "6":
            exportar_mes(DB_PATH, TEMPLATE_XLSX)
        elif op == "7":
            backfill_contadores(DB_PATH)
        elif op == "0":
            print("Bye.")
            break
        else:
            print("⚠️ Opción inválida.")

if __name__ == "__main__":
    # Permite ejecutar desde otra carpeta si el usuario copia el script:
    # - Busca plantilla en el mismo folder del script si no existe en cwd
    try:
        import pathlib
        here = pathlib.Path(__file__).resolve().parent
        if not os.path.exists(TEMPLATE_XLSX):
            cand = here / TEMPLATE_XLSX
            if cand.exists():
                TEMPLATE_XLSX = str(cand)
        if not os.path.exists(DB_PATH):
            cand = here / DB_PATH
            if cand.exists():
                DB_PATH = str(cand)
    except Exception:
        pass

    menu()
